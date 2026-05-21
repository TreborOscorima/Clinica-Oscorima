from __future__ import annotations

import json
import os
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, Iterable, List, Tuple

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from flask import current_app, has_app_context

SYSTEM_NAME = "WaykiSAC - Sistema de Gestion Clinica"
COMPANY_NAME = "Clinica Estetica OSCORIMA"
COMPANY_RUC = ""
COMPANY_ADDRESS = "Jiron las Begonias N° 29 Ayacucho-Peru"
COMPANY_PHONE = ""
COMPANY_VOUCHER = "BOLETA DE VENTA ELECTRONICA"
COMPANY_SERIE = "B001 - 000001"
EXPORT_DIR = Path(os.getenv("REPORT_EXPORT_DIR", "exports"))
HEADER_COLOR = "1F3C88"
ALT_ROW_COLOR = "F7F9FC"
GRID_COLOR = "D0D7ED"
DATE_FMT = "%d/%m/%Y %H:%M"

DEFAULT_HEADER_FIELDS = [
    ("Empresa", COMPANY_NAME),
    ("RUC", COMPANY_RUC or "-"),
    ("Direccion", COMPANY_ADDRESS),
    ("Telefono", COMPANY_PHONE or "-"),
    ("Tipo comprobante", COMPANY_VOUCHER),
    ("Serie y numero", COMPANY_SERIE),
]


class ExporterError(RuntimeError):
    """Custom exception for exporter failures."""


def _ensure_export_dir() -> Path:
    export_dir = Path(current_app.config.get("REPORT_EXPORT_DIR", EXPORT_DIR)) if has_app_context() else EXPORT_DIR
    export_dir.mkdir(parents=True, exist_ok=True)
    return export_dir


def _timestamp_path(prefix: str, ext: str) -> Path:
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M")
    safe_prefix = prefix.replace(" ", "_").lower()
    return _ensure_export_dir() / f"reporte_{safe_prefix}_{timestamp}{ext}"


def _normalize_filters(filters: Dict[str, Any] | None) -> Dict[str, Any]:
    filters = filters.copy() if isinstance(filters, dict) else {}
    meta = {
        "generated_at": filters.pop("generated_at", datetime.now(timezone.utc).strftime(DATE_FMT)),
        "generated_by": filters.pop("generated_by", "Sin usuario"),
        "date_range": filters.pop("date_range", "Sin rango"),
        "total_records": filters.pop("total_records", None),
        "summary": filters.pop("summary", []),
        "system_version": filters.pop("system_version", "v1"),
        "title": filters.pop("title", "Reporte"),
        "filters": filters.pop("filters", filters),
        "header_fields": filters.pop("header_fields", None),
        "include_logo_note": filters.pop("include_logo_note", True),
    }
    return meta



def _resolve_header_value(value: Any, meta: Dict[str, Any]) -> Any:
    if callable(value):
        try:
            return value(meta)
        except Exception:
            return value
    if isinstance(value, str):
        try:
            return value.format(**meta)
        except Exception:
            return value
    return value


def _default_header_rows(meta: Dict[str, Any]) -> List[Tuple[str, Any]]:
    rows: List[Tuple[str, Any]] = []
    for label, value in DEFAULT_HEADER_FIELDS:
        rows.append((label, _resolve_header_value(value, meta)))
    rows.extend([
        ("Reporte", _resolve_header_value("{title}", meta)),
        ("Fecha generacion", _resolve_header_value("{generated_at}", meta)),
        ("Generado por", _resolve_header_value("{generated_by}", meta)),
        ("Total registros", meta.get("total_records") if meta.get("total_records") is not None else "-"),
        ("Rango", _resolve_header_value("{date_range}", meta)),
    ])
    return rows

def _format_currency(value: Any) -> str:
    try:
        number = float(value or 0)
    except Exception:
        return str(value)
    return f"PEN {number:0.2f}"


def _apply_header_style(ws, header_row: int, columns: Iterable[int]) -> None:
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor=HEADER_COLOR)
    align_center = Alignment(horizontal="center", vertical="center")
    for col_idx in columns:
        cell = ws.cell(row=header_row, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center


def _apply_table_style(ws, header_row: int, column_count: int) -> None:
    thin_side = Side(border_style="thin", color=GRID_COLOR)
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    for row in ws.iter_rows(min_row=header_row, max_row=ws.max_row, min_col=1, max_col=column_count):
        for cell in row:
            cell.border = border
            if cell.row > header_row and cell.row % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=ALT_ROW_COLOR)
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00"
    ws.freeze_panes = ws[f"A{header_row + 1}"]
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(column_count)}{ws.max_row}"
    for col_idx in range(1, column_count + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 12
        for cell in ws[col_letter]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)


def _build_header_block(ws, meta: Dict[str, Any], column_count: int) -> int:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=column_count)
    title_cell = ws.cell(row=1, column=1, value=SYSTEM_NAME)
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor=HEADER_COLOR)
    title_cell.alignment = Alignment(horizontal="center")

    header_rows = meta.get("header_fields") or _default_header_rows(meta)
    row_cursor = 2
    for label, value in header_rows:
        if label is None:
            row_cursor += 1
            continue
        resolved = _resolve_header_value(value, meta)
        ws.cell(row=row_cursor, column=1, value=f"{label}:").font = Font(bold=True)
        ws.cell(row=row_cursor, column=2, value=resolved)
        row_cursor += 1

    ws.cell(row=row_cursor, column=1, value="Filtros:").font = Font(bold=True)
    filters = meta.get("filters") or {}
    if isinstance(filters, dict) and filters:
        for key, value in filters.items():
            ws.cell(row=row_cursor, column=2, value=f"{key}: {value}")
            row_cursor += 1
    elif isinstance(filters, str) and filters.strip():
        ws.cell(row=row_cursor, column=2, value=filters)
        row_cursor += 1
    else:
        ws.cell(row=row_cursor, column=2, value="Sin filtros")
        row_cursor += 1

    if meta.get("include_logo_note", True):
        ws.cell(row=row_cursor, column=1, value="Logo:").font = Font(italic=True)
        ws.cell(row=row_cursor, column=2, value="Espacio reservado para logo")
        row_cursor += 2
    else:
        row_cursor += 1

    return row_cursor


def export_to_excel(nombre_reporte: str, data: List[Dict[str, Any]], filters: Dict[str, Any] | None = None) -> str:
    meta = _normalize_filters(filters)
    if meta["total_records"] is None:
        meta["total_records"] = len(data)
    file_path = _timestamp_path(nombre_reporte, ".xlsx")

    workbook = Workbook()
    detail_ws = workbook.active
    detail_ws.title = "Detalle"

    columns = list(data[0].keys()) if data else []
    if not columns:
        columns = ["Mensaje"]
        data = [{"Mensaje": "Sin resultados en este periodo"}]

    start_row = _build_header_block(detail_ws, meta, len(columns))
    _apply_header_style(detail_ws, start_row, range(1, len(columns) + 1))
    for idx, field in enumerate(columns, start=1):
        detail_ws.cell(row=start_row, column=idx, value=field)

    for row_offset, item in enumerate(data, start=1):
        for col_idx, field in enumerate(columns, start=1):
            detail_ws.cell(row=start_row + row_offset, column=col_idx, value=item.get(field, ""))

    _apply_table_style(detail_ws, start_row, len(columns))

    summary_ws = workbook.create_sheet("Resumen")
    summary_rows: List[Tuple[str, Any]] = meta.get("summary") or []
    summary_ws.append(["Concepto", "Valor"])
    for label, value in summary_rows:
        summary_ws.append([label, value])
    for cell in summary_ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=HEADER_COLOR)
        cell.alignment = Alignment(horizontal="center")
    border_side = Side(border_style="thin", color=GRID_COLOR)
    for row in summary_ws.iter_rows(min_row=1, max_row=summary_ws.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    summary_ws.column_dimensions["A"].width = 40
    summary_ws.column_dimensions["B"].width = 25
    if len(summary_rows) >= 1:
        chart = BarChart()
        chart.title = "Resumen"
        chart.x_axis.title = "Concepto"
        chart.y_axis.title = "Valor"
        data_ref = Reference(summary_ws, min_col=2, min_row=1, max_row=summary_ws.max_row)
        cats_ref = Reference(summary_ws, min_col=1, min_row=2, max_row=summary_ws.max_row)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        summary_ws.add_chart(chart, "E2")

    audit_ws = workbook.create_sheet("Auditoria")
    audit_ws.append(["Campo", "Valor"])
    audit_entries = {
        "Sistema": SYSTEM_NAME,
        "Empresa": COMPANY_NAME,
        "Reporte": meta["title"],
        "Generado por": meta["generated_by"],
        "Fecha generacion": meta["generated_at"],
        "Total registros": meta["total_records"],
        "Version": meta["system_version"],
        "Rango": meta["date_range"],
        "Filtros": json.dumps(meta.get("filters"), ensure_ascii=False),
    }
    for key, value in audit_entries.items():
        audit_ws.append([key, value])
    for row in audit_ws.iter_rows(min_row=1, max_row=audit_ws.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    audit_ws.column_dimensions["A"].width = 28
    audit_ws.column_dimensions["B"].width = 60

    workbook.save(file_path)
    print(f"✔ Reporte generado: {file_path}")
    return str(file_path)


def export_to_pdf(nombre_reporte: str, data: List[Dict[str, Any]], filters: Dict[str, Any] | None = None) -> str:
    meta = _normalize_filters(filters)
    if meta["total_records"] is None:
        meta["total_records"] = len(data)
    file_path = _timestamp_path(nombre_reporte, ".pdf")

    doc = SimpleDocTemplate(
        str(file_path),
        pagesize=landscape(A4),
        leftMargin=36,
        rightMargin=36,
        topMargin=36,
        bottomMargin=36,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        name="ReportTitle",
        parent=styles["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=18,
        alignment=1,
        textColor=colors.HexColor(f"#{HEADER_COLOR}"),
    )
    normal = ParagraphStyle(name="Normal", parent=styles["Normal"], fontName="Helvetica", fontSize=10)
    bold = ParagraphStyle(name="Bold", parent=normal, fontName="Helvetica-Bold")

    story: List[Any] = []
    story.append(Paragraph(SYSTEM_NAME, title_style))
    story.append(Spacer(1, 8))
    header_rows = meta.get("header_fields") or _default_header_rows(meta)
    header_data = []
    for label, value in header_rows:
        if not label:
            continue
        resolved = _resolve_header_value(value, meta)
        header_data.append([Paragraph(f"<b>{label}</b>", normal), Paragraph(str(resolved), normal)])
    header_table = Table(header_data, colWidths=[160, 380])
    header_table.setStyle(TableStyle([
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(header_table)
    story.append(Spacer(1, 12))
    if meta.get("include_logo_note", True):
        story.append(Paragraph("Logo: espacio reservado", ParagraphStyle(name="LogoNote", parent=normal, italic=True)))
        story.append(Spacer(1, 12))

    columns = list(data[0].keys()) if data else []
    if not columns:
        columns = ["Mensaje"]
        data = [{"Mensaje": "Sin resultados en este periodo"}]

    table_data = [columns]
    for item in data:
        table_data.append([item.get(col, "") for col in columns])

    table = Table(table_data, repeatRows=1)
    style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{HEADER_COLOR}")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("TOPPADDING", (0, 0), (-1, 0), 6),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor(f"#{GRID_COLOR}")),
    ])
    for idx in range(1, len(table_data)):
        if idx % 2 == 0:
            style.add("BACKGROUND", (0, idx), (-1, idx), colors.HexColor(f"#{ALT_ROW_COLOR}"))
    table.setStyle(style)
    story.append(table)
    story.append(Spacer(1, 18))

    summary_rows = meta.get("summary") or []
    if summary_rows:
        story.append(Paragraph("Resumen", bold))
        summary_tbl = Table([[label, value] for label, value in summary_rows])
        summary_tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{HEADER_COLOR}")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor(f"#{GRID_COLOR}")),
        ]))
        story.append(summary_tbl)
        story.append(Spacer(1, 12))

    audit_json = json.dumps(meta.get("filters"), ensure_ascii=False)
    story.append(Paragraph(f"Filtros aplicados: {audit_json}", ParagraphStyle(name="Filters", parent=normal, fontSize=9)))
    story.append(Spacer(1, 6))
    story.append(Paragraph(f"Sistema: {SYSTEM_NAME} | Version: {meta['system_version']}", ParagraphStyle(name="Footer", parent=normal, fontSize=9)))

    doc.build(story)
    print(f"✔ Reporte generado: {file_path}")
    return str(file_path)
