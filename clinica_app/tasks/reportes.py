from __future__ import annotations

import os
from datetime import datetime, timezone
from typing import Any

from sqlalchemy import select
from sqlmodel import Session

from clinica_app.config import REPORT_EXPORT_DIR
from clinica_app.database import get_session
from clinica_app.models.caja import CajaMovimiento
from clinica_app.models.inventario import Compra, CompraItem, Producto, Proveedor
from clinica_app.models.paciente import Paciente
from clinica_app.models.turno import Turno


def generar_reporte(clinica_id: int, tipo: str, params: dict[str, Any]) -> str:
    """Genera el archivo Excel y retorna su path absoluto."""
    os.makedirs(REPORT_EXPORT_DIR, exist_ok=True)

    dispatch = {
        "pacientes":  _reporte_pacientes,
        "caja":       _reporte_caja,
        "turnos":     _reporte_turnos,
        "inventario": _reporte_inventario,
        "compras":    _reporte_compras,
    }
    fn = dispatch.get(tipo)
    if fn is None:
        raise ValueError(f"Tipo de reporte desconocido: {tipo}")
    return fn(clinica_id, params)


# ── Pacientes ─────────────────────────────────────────────────────────────────

def _reporte_pacientes(clinica_id: int, params: dict) -> str:
    import openpyxl

    with get_session() as session:
        rows = session.execute(
            select(Paciente)
            .where(Paciente.clinica_id == clinica_id, Paciente.is_active.is_(True))
            .order_by(Paciente.nombre)
        ).scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pacientes"
    ws.append(["ID", "Nombre", "Documento", "Email", "Teléfono",
                "Fecha Nac.", "Dirección", "Registrado"])
    for p in rows:
        ws.append([
            p.id, p.nombre, p.documento or "", p.email or "", p.telefono or "",
            p.fecha_nacimiento.strftime("%Y-%m-%d") if p.fecha_nacimiento else "",
            p.direccion or "",
            p.created_at.strftime("%Y-%m-%d") if p.created_at else "",
        ])

    return _guardar(wb, f"pacientes_{clinica_id}")


# ── Caja ──────────────────────────────────────────────────────────────────────

def _reporte_caja(clinica_id: int, params: dict) -> str:
    import openpyxl

    desde_str = params.get("desde")
    hasta_str = params.get("hasta")

    with get_session() as session:
        stmt = (
            select(CajaMovimiento)
            .where(CajaMovimiento.clinica_id == clinica_id, CajaMovimiento.is_active.is_(True))
        )
        if desde_str:
            stmt = stmt.where(CajaMovimiento.fecha >= datetime.fromisoformat(desde_str))
        if hasta_str:
            stmt = stmt.where(CajaMovimiento.fecha <= datetime.fromisoformat(hasta_str + "T23:59:59"))
        rows = session.execute(stmt.order_by(CajaMovimiento.fecha.desc())).scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Caja"
    ws.append(["ID", "Fecha", "Tipo", "Monto", "Método de Pago", "Observación"])
    for m in rows:
        ws.append([
            m.id,
            m.fecha.strftime("%Y-%m-%d %H:%M") if m.fecha else "",
            m.tipo.value if m.tipo else "",
            float(m.monto),
            m.metodo_pago.value if m.metodo_pago else "",
            m.observacion or "",
        ])

    return _guardar(wb, f"caja_{clinica_id}")


# ── Turnos ────────────────────────────────────────────────────────────────────

def _reporte_turnos(clinica_id: int, params: dict) -> str:
    import openpyxl

    desde_str = params.get("desde")
    hasta_str = params.get("hasta")

    with get_session() as session:
        stmt = (
            select(Turno)
            .where(Turno.clinica_id == clinica_id, Turno.is_active.is_(True))
        )
        if desde_str:
            stmt = stmt.where(Turno.fecha_hora >= datetime.fromisoformat(desde_str))
        if hasta_str:
            stmt = stmt.where(Turno.fecha_hora <= datetime.fromisoformat(hasta_str + "T23:59:59"))
        rows = session.execute(stmt.order_by(Turno.fecha_hora.desc())).scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Turnos"
    ws.append(["ID", "Paciente ID", "Profesional ID", "Servicio ID", "Fecha/Hora", "Estado"])
    for t in rows:
        ws.append([
            t.id, t.paciente_id, t.profesional_id or "", t.servicio_id or "",
            t.fecha_hora.strftime("%Y-%m-%d %H:%M") if t.fecha_hora else "",
            t.estado.value if t.estado else "",
        ])

    return _guardar(wb, f"turnos_{clinica_id}")


# ── Inventario ────────────────────────────────────────────────────────────────

def _reporte_inventario(clinica_id: int, params: dict) -> str:
    import openpyxl

    with get_session() as session:
        rows = session.execute(
            select(Producto)
            .where(Producto.clinica_id == clinica_id, Producto.is_active.is_(True))
            .order_by(Producto.nombre)
        ).scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario"
    ws.append(["ID", "SKU", "Nombre", "Stock Actual", "Stock Mínimo",
                "Precio Costo", "Precio Venta", "Activo"])
    for p in rows:
        ws.append([
            p.id, p.sku or "", p.nombre,
            float(p.stock_actual or 0), float(p.stock_minimo or 0),
            float(p.precio_costo or 0), float(p.precio_venta or 0),
            "Sí" if p.activo else "No",
        ])

    return _guardar(wb, f"inventario_{clinica_id}")


# ── Compras ───────────────────────────────────────────────────────────────────

def _reporte_compras(clinica_id: int, params: dict) -> str:
    import openpyxl

    desde_str = params.get("desde")
    hasta_str = params.get("hasta")

    with get_session() as session:
        stmt = (
            select(Compra, Proveedor)
            .outerjoin(Proveedor, Compra.proveedor_id == Proveedor.id)
            .where(Compra.clinica_id == clinica_id, Compra.is_active.is_(True))
        )
        if desde_str:
            stmt = stmt.where(Compra.fecha >= datetime.fromisoformat(desde_str))
        if hasta_str:
            stmt = stmt.where(Compra.fecha <= datetime.fromisoformat(hasta_str + "T23:59:59"))
        rows = session.execute(stmt.order_by(Compra.fecha.desc())).all()
        compra_rows = [(c, p) for c, p in rows]

        # Items: one sheet per compra would be complex; use flat detail sheet
        compra_ids = [c.id for c, _ in compra_rows]
        item_rows: list = []
        if compra_ids:
            items = session.execute(
                select(CompraItem, Producto)
                .join(Producto, CompraItem.producto_id == Producto.id)
                .where(CompraItem.compra_id.in_(compra_ids))
                .order_by(CompraItem.compra_id)
            ).all()
            item_rows = list(items)

    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "Compras"
    ws1.append(["ID", "Fecha", "Proveedor", "Tipo Doc", "Número",
                 "Nro Registro", "Total", "Observación"])
    for c, p in compra_rows:
        ws1.append([
            c.id,
            c.fecha.strftime("%Y-%m-%d") if c.fecha else "",
            p.nombre if p else "",
            c.tipo_doc or "", c.numero or "", c.nro_registro or "",
            float(c.total or 0), c.observacion or "",
        ])

    ws2 = wb.create_sheet("Detalle Items")
    ws2.append(["Compra ID", "Producto", "Cantidad", "Costo Unitario", "Subtotal"])
    for item, prod in item_rows:
        ws2.append([
            item.compra_id, prod.nombre,
            float(item.cantidad), float(item.costo_unitario), float(item.subtotal),
        ])

    return _guardar(wb, f"compras_{clinica_id}")


# ── Helpers ───────────────────────────────────────────────────────────────────

def _guardar(wb, prefix: str) -> str:
    ts = datetime.now(timezone.utc).strftime("%Y%m%d%H%M%S")
    filename = f"{prefix}_{ts}.xlsx"
    path = os.path.join(REPORT_EXPORT_DIR, filename)
    wb.save(path)
    return path
